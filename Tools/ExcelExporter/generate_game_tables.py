#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
游戏配置表生成工具
生成所有游戏所需的Excel配置表
"""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 项目根目录
PROJECT_ROOT = Path(__file__).parent.parent.parent
DATATABLES_DIR = PROJECT_ROOT / "DataTables"

# 样式定义
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_styled_workbook(headers, data, sheet_name="Sheet1"):
    """创建带样式的工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 写入表头
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    
    # 写入数据
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    # 中文字符宽度约为英文的2倍
                    chinese_count = sum(1 for c in str(cell.value) if '\u4e00' <= c <= '\u9fff')
                    cell_length += chinese_count
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    return wb


def generate_actor_table():
    """生成Actor配置表"""
    headers = [
        "Id", "Name", "ActorType", "MaxHealth", "Attack", "Defense", 
        "MoveSpeed", "CritRate", "CritDamage", "DodgeRate", 
        "AttackSpeed", "AttackRange", "FreezeResistance", 
        "PoisonResistance", "FireResistance", "PrefabPath", "Description"
    ]
    
    # ActorType: 1=Player, 2=Monster, 3=NPC
    data = [
        # 玩家角色
        [1, "新手玩家", 1, 100, 15, 10, 5.0, 0.1, 2.0, 0.05, 1.0, 15.0, 0.0, 0.0, 0.0, "Prefabs/Player/Player", "初始玩家角色"],
        [2, "老兵", 1, 150, 20, 15, 5.5, 0.15, 2.2, 0.08, 1.1, 15.0, 0.1, 0.1, 0.1, "Prefabs/Player/Veteran", "经验丰富的战士"],
        [3, "狙击手", 1, 80, 35, 5, 4.5, 0.25, 3.0, 0.1, 0.5, 30.0, 0.0, 0.0, 0.0, "Prefabs/Player/Sniper", "精准射击专家"],
        
        # 普通怪物
        [101, "变异僵尸", 2, 50, 8, 3, 3.0, 0.05, 1.5, 0.0, 0.8, 2.0, 0.0, 0.5, 0.0, "Prefabs/Monster/Zombie", "普通僵尸"],
        [102, "疯狗", 2, 30, 12, 2, 6.0, 0.1, 1.5, 0.15, 1.5, 1.5, 0.0, 0.0, 0.0, "Prefabs/Monster/Dog", "快速移动的疯狗"],
        [103, "武装匪徒", 2, 60, 15, 5, 4.0, 0.1, 1.8, 0.05, 1.0, 20.0, 0.0, 0.0, 0.0, "Prefabs/Monster/Bandit", "持枪的匪徒"],
        [104, "重装士兵", 2, 120, 10, 15, 2.5, 0.05, 1.5, 0.0, 0.6, 15.0, 0.2, 0.2, 0.2, "Prefabs/Monster/Heavy", "重装甲士兵"],
        
        # 精英怪物
        [201, "精英僵尸", 2, 150, 20, 8, 4.0, 0.1, 2.0, 0.05, 1.0, 2.5, 0.3, 0.6, 0.0, "Prefabs/Monster/EliteZombie", "强化僵尸"],
        [202, "狂暴犬", 2, 80, 25, 5, 7.0, 0.15, 2.0, 0.2, 2.0, 2.0, 0.0, 0.0, 0.3, "Prefabs/Monster/RageDog", "狂暴状态的犬类"],
        [203, "佣兵队长", 2, 200, 25, 12, 4.5, 0.15, 2.2, 0.1, 1.2, 25.0, 0.1, 0.1, 0.1, "Prefabs/Monster/MercLeader", "佣兵队长"],
        
        # Boss
        [301, "尸潮之王", 2, 1000, 35, 20, 3.0, 0.15, 2.5, 0.0, 0.8, 3.0, 0.5, 0.8, 0.3, "Prefabs/Monster/ZombieKing", "僵尸Boss"],
        [302, "机甲守卫", 2, 2000, 50, 40, 2.0, 0.1, 2.0, 0.0, 0.5, 20.0, 0.6, 0.6, 0.6, "Prefabs/Monster/MechGuard", "机械Boss"],
        
        # NPC
        [501, "商人老王", 3, 100, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "Prefabs/NPC/Merchant", "武器商人"],
        [502, "任务员小李", 3, 100, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "Prefabs/NPC/QuestGiver", "任务发布员"],
        [503, "铁匠张", 3, 100, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "Prefabs/NPC/Blacksmith", "装备强化师"],
        [504, "医疗兵", 3, 100, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "Prefabs/NPC/Medic", "可以治疗玩家"],
    ]
    
    wb = create_styled_workbook(headers, data, "Actor")
    return wb


def generate_monster_table():
    """生成怪物配置表"""
    headers = [
        "Id", "ActorId", "MonsterType", "ExpReward", "GoldReward", 
        "DetectionRange", "ChaseRange", "CanFlee", "FleeHealthPercent",
        "DropItemIds", "DropRates", "PatrolRadius"
    ]
    
    # MonsterType: 0=Normal, 1=Elite, 2=Boss, 3=Minion
    data = [
        # 普通怪物
        [1, 101, 0, 30, 10, 12.0, 20.0, False, 0.0, "1001,1002", "0.3,0.1", 10.0],
        [2, 102, 0, 25, 8, 15.0, 25.0, True, 0.2, "1001", "0.2", 15.0],
        [3, 103, 0, 40, 15, 20.0, 30.0, True, 0.3, "1002,2001", "0.2,0.05", 8.0],
        [4, 104, 0, 50, 20, 15.0, 25.0, False, 0.0, "1002,2002", "0.25,0.08", 5.0],
        
        # 精英怪物
        [101, 201, 1, 100, 50, 15.0, 30.0, False, 0.0, "1002,2001,2002", "0.5,0.2,0.15", 12.0],
        [102, 202, 1, 80, 40, 18.0, 35.0, True, 0.15, "1001,2001", "0.4,0.25", 20.0],
        [103, 203, 1, 150, 80, 25.0, 40.0, True, 0.2, "2002,2003,3001", "0.4,0.3,0.1", 10.0],
        
        # Boss
        [201, 301, 2, 500, 300, 20.0, 50.0, False, 0.0, "2003,3001,3002", "0.8,0.5,0.3", 0.0],
        [202, 302, 2, 1000, 500, 25.0, 60.0, False, 0.0, "3001,3002,3003", "0.9,0.6,0.4", 0.0],
    ]
    
    wb = create_styled_workbook(headers, data, "Monster")
    return wb


def generate_npc_table():
    """生成NPC配置表"""
    headers = [
        "Id", "ActorId", "NPCType", "Dialogues", "ShopItemIds", 
        "QuestIds", "InteractionRange", "IsStationary"
    ]
    
    # NPCType: 0=Normal, 1=Merchant, 2=QuestGiver, 3=Trainer, 4=Blacksmith, 5=Teleporter
    data = [
        [1, 501, 1, "欢迎光临!|需要点什么?|祝你好运!", "2001,2002,2003,1001,1002", "", 3.0, True],
        [2, 502, 2, "英雄,我需要你的帮助!|请帮我消灭那些怪物!", "", "1,2,3", 3.0, True],
        [3, 503, 4, "要强化装备吗?|好材料才能打造好武器!", "", "", 3.0, True],
        [4, 504, 0, "需要治疗吗?|保重身体!", "", "", 3.0, False],
    ]
    
    wb = create_styled_workbook(headers, data, "NPC")
    return wb


def generate_item_table():
    """生成物品基础配置表"""
    headers = [
        "Id", "Name", "ItemType", "Quality", "Stackable", "MaxStack",
        "BuyPrice", "SellPrice", "IconPath", "Description"
    ]
    
    # ItemType: 0=Consumable, 1=Equipment, 2=Weapon, 3=Ammo, 4=Material, 5=Quest
    # Quality: 0=Common, 1=Uncommon, 2=Rare, 3=Epic, 4=Legendary
    data = [
        # 消耗品
        [1001, "小型治疗包", 0, 0, True, 10, 50, 25, "Icons/Item/HealSmall", "恢复30点生命"],
        [1002, "大型治疗包", 0, 1, True, 5, 150, 75, "Icons/Item/HealLarge", "恢复100点生命"],
        [1003, "急救箱", 0, 2, True, 3, 400, 200, "Icons/Item/MedKit", "恢复50%最大生命"],
        [1004, "解毒剂", 0, 0, True, 10, 80, 40, "Icons/Item/Antidote", "清除所有负面效果"],
        [1005, "兴奋剂", 0, 1, True, 5, 200, 100, "Icons/Item/Stimulant", "提升30%移动速度持续10秒"],
        [1006, "护盾注射器", 0, 2, True, 3, 350, 175, "Icons/Item/Shield", "获得50点护盾持续15秒"],
        
        # 弹药
        [1101, "手枪弹药", 3, 0, True, 100, 10, 5, "Icons/Item/AmmoPistol", "手枪用弹药"],
        [1102, "步枪弹药", 3, 0, True, 120, 15, 7, "Icons/Item/AmmoRifle", "步枪用弹药"],
        [1103, "霰弹", 3, 0, True, 40, 20, 10, "Icons/Item/AmmoShotgun", "霰弹枪用弹药"],
        [1104, "狙击弹", 3, 1, True, 20, 30, 15, "Icons/Item/AmmoSniper", "狙击枪用弹药"],
        
        # 材料
        [1201, "破损零件", 4, 0, True, 99, 20, 10, "Icons/Item/PartBroken", "可用于修理装备"],
        [1202, "精密零件", 4, 1, True, 50, 100, 50, "Icons/Item/PartPrecise", "高品质零件"],
        [1203, "能量核心", 4, 2, True, 20, 500, 250, "Icons/Item/EnergyCore", "蕴含能量的核心"],
        
        # 任务物品
        [1301, "神秘钥匙", 5, 2, False, 1, 0, 0, "Icons/Item/MysteryKey", "开启神秘房间的钥匙"],
        [1302, "情报文件", 5, 1, False, 1, 0, 0, "Icons/Item/Intel", "重要的情报文件"],
    ]
    
    wb = create_styled_workbook(headers, data, "Item")
    return wb


def generate_weapon_table():
    """生成武器配置表"""
    headers = [
        "Id", "Name", "WeaponType", "ElementType", "Quality", 
        "Damage", "FireRate", "Range", "MagazineSize", "ReloadTime",
        "Spread", "Recoil", "Penetration", "PelletCount",
        "HealthBonus", "AttackBonus", "DefenseBonus", "CritRateBonus",
        "RequiredLevel", "BuyPrice", "SellPrice", "PrefabPath", "Description"
    ]
    
    # WeaponType: 0=Pistol, 1=Rifle, 2=SMG, 3=Shotgun, 4=Sniper, 5=MachineGun, 6=Melee
    # ElementType: 0=None, 1=Ice, 2=Poison, 3=Fire, 4=Lightning
    data = [
        # 手枪
        [2001, "M9手枪", 0, 0, 0, 15, 5.0, 25.0, 15, 1.5, 2.0, 0.5, 0, 1, 0, 0, 0, 0, 1, 200, 100, "Prefabs/Weapon/M9", "标准手枪"],
        [2002, "沙漠之鹰", 0, 0, 2, 35, 2.5, 30.0, 7, 2.0, 3.0, 1.5, 1, 1, 0, 5, 0, 0.02, 5, 800, 400, "Prefabs/Weapon/DesertEagle", "威力强大的手枪"],
        [2003, "冰霜左轮", 0, 1, 3, 28, 1.5, 28.0, 6, 2.5, 2.5, 2.0, 0, 1, 0, 3, 0, 0.05, 10, 1500, 750, "Prefabs/Weapon/FrostRevolver", "附带冰冻效果"],
        
        # 步枪
        [2101, "M4A1突击步枪", 1, 0, 1, 20, 10.0, 40.0, 30, 2.0, 2.0, 1.0, 0, 1, 0, 3, 0, 0.02, 3, 600, 300, "Prefabs/Weapon/M4A1", "平衡性良好的突击步枪"],
        [2102, "AK-47", 1, 0, 1, 25, 8.0, 35.0, 30, 2.5, 3.0, 1.5, 1, 1, 0, 5, 0, 0.03, 5, 800, 400, "Prefabs/Weapon/AK47", "伤害高但后坐力大"],
        [2103, "燃烧突击步枪", 1, 3, 3, 22, 9.0, 38.0, 25, 2.2, 2.5, 1.2, 0, 1, 0, 8, 0, 0.05, 15, 2500, 1250, "Prefabs/Weapon/FireRifle", "子弹附带燃烧效果"],
        
        # 冲锋枪
        [2201, "MP5", 2, 0, 0, 12, 15.0, 25.0, 30, 1.8, 3.5, 0.8, 0, 1, 0, 0, 0, 0.01, 1, 400, 200, "Prefabs/Weapon/MP5", "射速快的冲锋枪"],
        [2202, "毒蛇冲锋枪", 2, 2, 2, 14, 12.0, 28.0, 35, 2.0, 3.0, 0.9, 0, 1, 0, 2, 0, 0.03, 8, 1200, 600, "Prefabs/Weapon/ViperSMG", "子弹带有剧毒"],
        
        # 霰弹枪
        [2301, "M870霰弹枪", 3, 0, 1, 8, 1.2, 15.0, 8, 3.0, 15.0, 2.0, 0, 8, 0, 0, 5, 0, 3, 700, 350, "Prefabs/Weapon/M870", "近距离威力巨大"],
        [2302, "双管猎枪", 3, 0, 2, 12, 0.8, 12.0, 2, 2.5, 18.0, 2.5, 0, 10, 0, 5, 0, 0.05, 8, 1000, 500, "Prefabs/Weapon/DoubleBarrel", "单次伤害极高"],
        [2303, "雷霆霰弹枪", 3, 4, 3, 10, 1.0, 18.0, 6, 3.5, 12.0, 2.0, 0, 8, 0, 8, 3, 0.08, 20, 3000, 1500, "Prefabs/Weapon/ThunderShotgun", "雷电霰弹"],
        
        # 狙击枪
        [2401, "M24狙击枪", 4, 0, 2, 80, 0.8, 80.0, 5, 3.0, 0.2, 2.5, 2, 1, 0, 10, 0, 0.15, 10, 1500, 750, "Prefabs/Weapon/M24", "精准狙击"],
        [2402, "巴雷特", 4, 0, 3, 150, 0.5, 100.0, 5, 4.0, 0.1, 3.5, 3, 1, 0, 25, 0, 0.2, 20, 5000, 2500, "Prefabs/Weapon/Barrett", "反器材狙击枪"],
        
        # 机枪
        [2501, "M249机枪", 5, 0, 2, 18, 12.0, 45.0, 100, 5.0, 5.0, 1.8, 1, 1, 20, 5, 0, 0.02, 12, 2000, 1000, "Prefabs/Weapon/M249", "弹容量大的机枪"],
        
        # 近战武器
        [2601, "战术匕首", 6, 0, 0, 30, 2.0, 2.0, 1, 0, 0, 0, 0, 1, 0, 0, 0, 0.1, 1, 150, 75, "Prefabs/Weapon/Knife", "快速近战武器"],
        [2602, "消防斧", 6, 0, 1, 60, 1.0, 2.5, 1, 0, 0, 0, 0, 1, 0, 10, 5, 0.05, 5, 500, 250, "Prefabs/Weapon/FireAxe", "重型近战武器"],
    ]
    
    wb = create_styled_workbook(headers, data, "Weapon")
    return wb


def generate_equipment_table():
    """生成装备配置表"""
    headers = [
        "Id", "Name", "Slot", "Quality", "RequiredLevel",
        "HealthBonus", "AttackBonus", "DefenseBonus", 
        "CritRateBonus", "CritDamageBonus", "DodgeRateBonus",
        "MoveSpeedBonus", "AttackSpeedBonus",
        "FreezeResistance", "PoisonResistance", "FireResistance",
        "BuyPrice", "SellPrice", "PrefabPath", "Description"
    ]
    
    # Slot: 0=Head, 1=Chest, 2=Gloves, 3=Legs, 4=Boots, 5=PrimaryWeapon, 6=SecondaryWeapon, 7=Backpack, 8=Accessory1, 9=Accessory2
    data = [
        # 头盔
        [3001, "战术头盔", 0, 0, 1, 10, 0, 5, 0, 0, 0, 0, 0, 0, 0, 0, 200, 100, "Prefabs/Equip/TacticalHelmet", "基础防护头盔"],
        [3002, "防弹头盔", 0, 1, 5, 20, 0, 10, 0, 0, 0.02, 0, 0, 0, 0, 0, 500, 250, "Prefabs/Equip/BulletproofHelmet", "军用防弹头盔"],
        [3003, "精英头盔", 0, 2, 10, 35, 0, 18, 0.02, 0, 0.03, 0, 0, 0.1, 0.1, 0.1, 1200, 600, "Prefabs/Equip/EliteHelmet", "精英士兵装备"],
        [3004, "传奇头盔", 0, 4, 20, 60, 5, 30, 0.05, 0.1, 0.05, 0, 0, 0.2, 0.2, 0.2, 5000, 2500, "Prefabs/Equip/LegendHelmet", "传奇级头盔"],
        
        # 护甲
        [3101, "战术背心", 1, 0, 1, 20, 0, 8, 0, 0, 0, 0, 0, 0, 0, 0, 300, 150, "Prefabs/Equip/TacticalVest", "基础护甲"],
        [3102, "防弹衣", 1, 1, 5, 40, 0, 18, 0, 0, 0, -0.2, 0, 0, 0, 0, 800, 400, "Prefabs/Equip/BulletproofVest", "防弹背心"],
        [3103, "重型装甲", 1, 2, 10, 70, 0, 35, 0, 0, 0, -0.5, 0, 0.15, 0.15, 0.15, 2000, 1000, "Prefabs/Equip/HeavyArmor", "重型防护装甲"],
        [3104, "外骨骼装甲", 1, 3, 15, 100, 10, 45, 0.03, 0.1, 0.02, 0, 0.1, 0.2, 0.2, 0.2, 4000, 2000, "Prefabs/Equip/ExoArmor", "高科技外骨骼"],
        
        # 手套
        [3201, "战术手套", 2, 0, 1, 0, 3, 2, 0.01, 0, 0, 0, 0.05, 0, 0, 0, 150, 75, "Prefabs/Equip/TacticalGloves", "提升射击精度"],
        [3202, "射击手套", 2, 1, 5, 0, 5, 3, 0.03, 0.05, 0, 0, 0.08, 0, 0, 0, 400, 200, "Prefabs/Equip/ShootingGloves", "专业射击手套"],
        [3203, "精密手套", 2, 2, 10, 0, 8, 5, 0.05, 0.1, 0, 0, 0.12, 0, 0, 0, 1000, 500, "Prefabs/Equip/PrecisionGloves", "大幅提升精准度"],
        
        # 腿甲
        [3301, "战术护腿", 3, 0, 1, 15, 0, 5, 0, 0, 0.01, 0.2, 0, 0, 0, 0, 250, 125, "Prefabs/Equip/TacticalLegs", "基础护腿"],
        [3302, "强化护腿", 3, 1, 5, 25, 0, 10, 0, 0, 0.02, 0.3, 0, 0, 0, 0, 600, 300, "Prefabs/Equip/ReinforcedLegs", "强化护腿"],
        
        # 靴子
        [3401, "战术靴", 4, 0, 1, 5, 0, 3, 0, 0, 0.02, 0.5, 0, 0, 0, 0, 200, 100, "Prefabs/Equip/TacticalBoots", "提升移动速度"],
        [3402, "急行靴", 4, 1, 5, 10, 0, 5, 0, 0, 0.04, 1.0, 0, 0, 0, 0, 500, 250, "Prefabs/Equip/SprintBoots", "大幅提升移动速度"],
        [3403, "隐匿靴", 4, 2, 10, 15, 0, 8, 0, 0, 0.08, 0.8, 0, 0, 0, 0, 1200, 600, "Prefabs/Equip/StealthBoots", "提升闪避率"],
        
        # 背包
        [3501, "小型背包", 7, 0, 1, 0, 0, 0, 0, 0, 0, -0.1, 0, 0, 0, 0, 100, 50, "Prefabs/Equip/SmallBackpack", "+5格背包容量"],
        [3502, "中型背包", 7, 1, 5, 0, 0, 0, 0, 0, 0, -0.2, 0, 0, 0, 0, 300, 150, "Prefabs/Equip/MediumBackpack", "+10格背包容量"],
        [3503, "大型背包", 7, 2, 10, 0, 0, 0, 0, 0, 0, -0.3, 0, 0, 0, 0, 800, 400, "Prefabs/Equip/LargeBackpack", "+20格背包容量"],
        
        # 饰品
        [3601, "幸运吊坠", 8, 1, 5, 0, 0, 0, 0.05, 0.1, 0.02, 0, 0, 0, 0, 0, 500, 250, "Prefabs/Equip/LuckyPendant", "提升暴击和闪避"],
        [3602, "力量戒指", 9, 1, 5, 0, 8, 0, 0.02, 0.05, 0, 0, 0.05, 0, 0, 0, 600, 300, "Prefabs/Equip/PowerRing", "提升攻击力"],
        [3603, "生命护符", 8, 2, 10, 50, 0, 5, 0, 0, 0, 0, 0, 0.1, 0.1, 0.1, 1500, 750, "Prefabs/Equip/LifeAmulet", "提升生命和抗性"],
    ]
    
    wb = create_styled_workbook(headers, data, "Equipment")
    return wb


def generate_consumable_table():
    """生成消耗品配置表"""
    headers = [
        "Id", "ItemId", "EffectType", "EffectValue", "Duration",
        "BuffType", "Cooldown", "Description"
    ]
    
    # EffectType: 0=HealHealth, 1=HealHealthPercent, 2=AddBuff, 3=RemoveDebuff, 4=AddAmmo, 5=AddExperience
    # BuffType: 0=None, 1=Freeze, 2=Poison, 3=Burn, 4=Stun, 5=Speed, 6=Shield, 7=Heal, 8=AttackBuff, 9=DefenseBuff
    data = [
        [1, 1001, 0, 30, 0, 0, 1.0, "立即恢复30点生命值"],
        [2, 1002, 0, 100, 0, 0, 1.5, "立即恢复100点生命值"],
        [3, 1003, 1, 0.5, 0, 0, 3.0, "立即恢复50%最大生命值"],
        [4, 1004, 3, 0, 0, 0, 2.0, "清除所有负面状态"],
        [5, 1005, 2, 0.3, 10, 5, 30.0, "提升30%移动速度，持续10秒"],
        [6, 1006, 2, 50, 15, 6, 45.0, "获得50点护盾，持续15秒"],
        
        # 弹药转换为使用效果
        [101, 1101, 4, 30, 0, 0, 0, "补充30发手枪弹药"],
        [102, 1102, 4, 60, 0, 0, 0, "补充60发步枪弹药"],
        [103, 1103, 4, 16, 0, 0, 0, "补充16发霰弹"],
        [104, 1104, 4, 10, 0, 0, 0, "补充10发狙击弹"],
    ]
    
    wb = create_styled_workbook(headers, data, "Consumable")
    return wb


def generate_buff_table():
    """生成Buff配置表"""
    headers = [
        "Id", "Name", "BuffType", "DefaultValue", "DefaultDuration",
        "MaxStack", "IconPath", "EffectPath", "Description"
    ]
    
    # BuffType: 0=None, 1=Freeze, 2=Poison, 3=Burn, 4=Stun, 5=Speed, 6=Shield, 7=Heal, 8=AttackBuff, 9=DefenseBuff
    data = [
        [1, "冰冻", 1, 0.3, 3.0, 5, "Icons/Buff/Freeze", "Effects/Freeze", "减速30%每层,可叠加5层"],
        [2, "中毒", 2, 5.0, 5.0, 5, "Icons/Buff/Poison", "Effects/Poison", "每秒5点伤害每层,可叠加5层"],
        [3, "燃烧", 3, 8.0, 4.0, 3, "Icons/Buff/Burn", "Effects/Burn", "每秒8点伤害,增加受到的伤害"],
        [4, "眩晕", 4, 0, 1.0, 1, "Icons/Buff/Stun", "Effects/Stun", "无法行动"],
        [5, "加速", 5, 0.3, 10.0, 1, "Icons/Buff/Speed", "Effects/Speed", "移动速度+30%"],
        [6, "护盾", 6, 50.0, 15.0, 1, "Icons/Buff/Shield", "Effects/Shield", "吸收50点伤害"],
        [7, "治疗", 7, 5.0, 10.0, 1, "Icons/Buff/Heal", "Effects/Heal", "每秒恢复5点生命"],
        [8, "攻击强化", 8, 0.2, 15.0, 1, "Icons/Buff/AttackUp", "Effects/AttackUp", "攻击力+20%"],
        [9, "防御强化", 9, 0.2, 15.0, 1, "Icons/Buff/DefenseUp", "Effects/DefenseUp", "防御力+20%"],
    ]
    
    wb = create_styled_workbook(headers, data, "Buff")
    return wb


def generate_quest_table():
    """生成任务配置表"""
    headers = [
        "Id", "Name", "NPCId", "Type", "TargetType", "TargetId",
        "TargetCount", "RewardExp", "RewardGold", "RewardItemIds",
        "RewardItemCounts", "PreQuestId", "Description"
    ]
    
    # Type: 0=Kill, 1=Collect, 2=Talk, 3=Explore
    data = [
        [1, "清理僵尸", 2, 0, 0, 101, 10, 100, 50, "1001,1101", "3,30", 0, "消灭10只变异僵尸"],
        [2, "收集零件", 2, 1, 0, 1201, 5, 80, 30, "1201", "2", 0, "收集5个破损零件"],
        [3, "精英猎杀", 2, 0, 0, 201, 3, 300, 150, "2001,1002", "1,2", 1, "消灭3只精英僵尸"],
        [4, "与铁匠对话", 2, 2, 0, 503, 1, 50, 20, "", "", 2, "与铁匠张交谈"],
        [5, "Boss讨伐", 2, 0, 0, 301, 1, 1000, 500, "3001,1003", "1,5", 3, "击败尸潮之王"],
    ]
    
    wb = create_styled_workbook(headers, data, "Quest")
    return wb


def generate_spawn_point_table():
    """生成刷怪点配置表"""
    headers = [
        "Id", "MonsterId", "PosX", "PosY", "PosZ", 
        "RespawnTime", "MaxCount", "PatrolRadius", "Description"
    ]
    
    data = [
        [1, 1, 20.0, 0.0, 20.0, 30.0, 3, 10.0, "东部僵尸刷新点"],
        [2, 1, -20.0, 0.0, 20.0, 30.0, 3, 10.0, "西部僵尸刷新点"],
        [3, 2, 15.0, 0.0, -15.0, 25.0, 2, 15.0, "疯狗刷新点"],
        [4, 3, 0.0, 0.0, 30.0, 45.0, 2, 8.0, "匪徒刷新点"],
        [5, 101, 30.0, 0.0, 0.0, 120.0, 1, 12.0, "精英僵尸刷新点"],
        [6, 102, -30.0, 0.0, 0.0, 120.0, 1, 20.0, "狂暴犬刷新点"],
        [7, 201, 0.0, 0.0, 50.0, 600.0, 1, 0.0, "Boss刷新点"],
    ]
    
    wb = create_styled_workbook(headers, data, "SpawnPoint")
    return wb


def main():
    """主函数"""
    print("=" * 60)
    print("游戏配置表生成工具")
    print("=" * 60)
    
    # 确保目录存在
    DATATABLES_DIR.mkdir(parents=True, exist_ok=True)
    
    # 生成所有配置表
    tables = [
        ("Actor", generate_actor_table),
        ("Monster", generate_monster_table),
        ("NPC", generate_npc_table),
        ("Item", generate_item_table),
        ("Weapon", generate_weapon_table),
        ("Equipment", generate_equipment_table),
        ("Consumable", generate_consumable_table),
        ("Buff", generate_buff_table),
        ("Quest", generate_quest_table),
        ("SpawnPoint", generate_spawn_point_table),
    ]
    
    for table_name, generator in tables:
        wb = generator()
        file_path = DATATABLES_DIR / f"{table_name}.xlsx"
        wb.save(file_path)
        print(f"已生成: {file_path}")
    
    print()
    print("=" * 60)
    print("配置表生成完成！")
    print("=" * 60)
    print(f"文件位置: {DATATABLES_DIR}")
    print()
    print("下一步:")
    print("  1. 运行 python excel_to_unity.py 导出数据到Unity")
    print("  2. 或运行 export.bat (Windows) / export.sh (Linux/Mac)")


if __name__ == "__main__":
    main()

