#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel数据导出工具
将DataTables目录下的.xlsx文件导出为Unity可用的格式
"""

import os
import json
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl import load_workbook

# 项目根目录（现在在Tools/ExcelExporter子目录中，需要向上3级）
PROJECT_ROOT = Path(__file__).parent.parent.parent
DATATABLES_DIR = PROJECT_ROOT / "DataTables"
UNITY_RESOURCES_DIR = PROJECT_ROOT / "GamePro" / "Assets" / "Resources" / "ExcelImporter"
UNITY_CODE_DIR = PROJECT_ROOT / "GamePro" / "Assets" / "Code" / "ExcelImporter"


def ensure_directories():
    """确保输出目录存在"""
    UNITY_RESOURCES_DIR.mkdir(parents=True, exist_ok=True)
    UNITY_CODE_DIR.mkdir(parents=True, exist_ok=True)


def to_csharp_type(excel_type: str, value: Any) -> str:
    """将Excel数据类型转换为C#类型"""
    if value is None:
        return "string"
    
    # 如果已经是Python类型，直接判断
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    
    value_str = str(value).strip()
    
    # 空字符串默认为string
    if not value_str:
        return "string"
    
    # 尝试判断类型
    if value_str.lower() in ('true', 'false'):
        return "bool"
    
    # 尝试解析为整数
    try:
        int(value_str)
        return "int"
    except ValueError:
        pass
    
    # 尝试解析为浮点数
    try:
        float(value_str)
        return "float"
    except ValueError:
        pass
    
    return "string"


def parse_excel_file(file_path: Path) -> Dict[str, Any]:
    """解析Excel文件，返回表名和数据"""
    print(f"正在解析文件: {file_path.name}")
    
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"错误: 无法加载文件 {file_path.name}: {e}")
        return None
    
    result = {}
    
    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 跳过空表
        if ws.max_row < 2:
            continue
        
        # 读取表头（第一行）
        headers = []
        for cell in ws[1]:
            header = str(cell.value).strip() if cell.value else f"Column{cell.column}"
            headers.append(header)
        
        # 读取数据行
        data_rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # 处理不同类型的值
                if value is None:
                    row_data[header] = None
                elif isinstance(value, (int, float)):
                    row_data[header] = value
                elif isinstance(value, bool):
                    row_data[header] = value
                else:
                    row_data[header] = str(value).strip()
            
            # 跳过完全空的行
            if any(v is not None and str(v).strip() != "" for v in row_data.values()):
                data_rows.append(row_data)
        
        if data_rows:
            result[sheet_name] = {
                "headers": headers,
                "data": data_rows
            }
    
    return result if result else None


def generate_csharp_class(table_name: str, headers: List[str], data_rows: List[Dict[str, Any]]) -> str:
    """生成C#数据类代码"""
    class_name = to_pascal_case(table_name)
    
    # 推断字段类型（基于所有数据行，而不仅仅是第一行）
    field_types = {}
    for header in headers:
        # 遍历所有数据行，找到第一个非空值来推断类型
        inferred_type = "string"  # 默认类型
        for row in data_rows:
            value = row.get(header)
            if value is not None and str(value).strip():
                inferred_type = to_csharp_type(header, value)
                # 如果推断为int，检查是否所有值都是int，否则可能是float
                if inferred_type == "int":
                    # 检查后续值，如果有float则升级为float
                    for check_row in data_rows:
                        check_value = check_row.get(header)
                        if check_value is not None:
                            try:
                                float(str(check_value))
                                if "." in str(check_value):
                                    inferred_type = "float"
                                    break
                            except:
                                pass
                break
        field_types[header] = inferred_type
    
    # 生成类代码
    code = f"""using System;

namespace ExcelImporter
{{
    /// <summary>
    /// {table_name}数据表
    /// </summary>
    [Serializable]
    public class {class_name}Data
    {{
"""
    
    # 生成字段（使用属性，Unity JsonUtility需要public字段或属性）
    for header in headers:
        csharp_type = field_types[header]
        property_name = to_pascal_case(header)
        # Unity JsonUtility需要public字段，但也可以使用属性
        # 为了更好的兼容性，使用public字段
        field_name = to_camel_case(header)
        code += f"        public {csharp_type} {property_name};\n"
    
    code += """    }
}
"""
    
    return code


def generate_data_manager_code(tables: Dict[str, Dict[str, Any]]) -> str:
    """生成数据管理器代码"""
    code = """using System;
using System.Collections.Generic;
using System.Linq;
using System.Reflection;
using System.Text;
using UnityEngine;

namespace ExcelImporter
{
    /// <summary>
    /// Excel数据管理器 - 一次性加载所有数据
    /// </summary>
    public static class ExcelDataManager
    {
        private static bool _isInitialized = false;
        private const BindingFlags FieldFlags = BindingFlags.Public | BindingFlags.Instance;
"""
    
    # 为每个表生成数据字典和初始化代码
    for table_name, table_info in tables.items():
        class_name = to_pascal_case(table_name)
        dict_name = f"_{to_camel_case(table_name)}Dict"
        list_name = f"_{to_camel_case(table_name)}List"
        
        code += f"""
        private static Dictionary<int, {class_name}Data> {dict_name} = new Dictionary<int, {class_name}Data>();
        private static List<{class_name}Data> {list_name} = new List<{class_name}Data>();
"""
    
    # 初始化方法
    code += """
        /// <summary>
        /// 初始化并加载所有数据（只需调用一次）
        /// </summary>
        public static void Initialize()
        {
            if (_isInitialized)
            {
                Debug.LogWarning("ExcelDataManager已经初始化过了");
                return;
            }
"""
    
    for table_name, table_info in tables.items():
        class_name = to_pascal_case(table_name)
        dict_name = f"_{to_camel_case(table_name)}Dict"
        list_name = f"_{to_camel_case(table_name)}List"
        
        code += f"""
            // 加载{table_name}数据
            var {to_camel_case(table_name)}Json = Resources.Load<TextAsset>("ExcelImporter/{table_name}");
            if ({to_camel_case(table_name)}Json != null && !string.IsNullOrEmpty({to_camel_case(table_name)}Json.text))
            {{
                try
                {{
                    // Unity JsonUtility要求JSON格式必须严格匹配类结构
                    // 确保JSON字符串格式正确
                    var jsonText = {to_camel_case(table_name)}Json.text.Trim();
                    if (string.IsNullOrEmpty(jsonText))
                    {{
                        Debug.LogError($"加载{table_name}数据失败: JSON内容为空");
                        return;
                    }}
                    
                    // Unity JsonUtility对数组反序列化有已知问题，使用包装类方式
                    // 注意：JSON格式必须是 {{"items":[...]}} 格式
                    var {to_camel_case(table_name)}Array = JsonUtility.FromJson<{class_name}DataArray>(jsonText);
                    
                    if ({to_camel_case(table_name)}Array == null)
                    {{
                        Debug.LogError($"加载{table_name}数据失败: 反序列化结果为null");
                        Debug.LogError($"JSON前200字符: {{jsonText.Substring(0, Math.Min(200, jsonText.Length))}}");
                        return;
                    }}
                    
                    if ({to_camel_case(table_name)}Array.items == null)
                    {{
                        Debug.LogWarning($"加载{table_name}数据失败: items数组为null");
                        Debug.LogWarning($"JSON前200字符: {{jsonText.Substring(0, Math.Min(200, jsonText.Length))}}");
                        // 尝试手动解析JSON（备用方案）
                        Debug.LogWarning("尝试备用解析方案...");
                        return;
                    }}
                    
                    Debug.Log($"加载{table_name}数据: {{ {to_camel_case(table_name)}Array.items.Length }} 条记录");
                    
                    int loadedCount = 0;
                    int nullCount = 0;
                    foreach (var item in {to_camel_case(table_name)}Array.items)
                    {{
                        if (item != null)
                        {{
                            {list_name}.Add(item);
                            loadedCount++;
                            
                            // 尝试使用Id字段，如果没有则使用第一个字段
                            var idValue = GetIdValue(item);
                            if (idValue != null && idValue != 0)
                            {{
                                {dict_name}[idValue.Value] = item;
                            }}
                        }}
                        else
                        {{
                            nullCount++;
                        }}
                    }}
                    
                    if (nullCount > 0)
                    {{
                        Debug.LogWarning($"加载{table_name}数据时发现 {{nullCount}} 个null项");
                    }}
                    
                    Debug.Log($"成功加载{table_name}数据: {{loadedCount}} 条记录到列表，{{ {dict_name}.Count }} 条记录到字典");
                }}
                catch (System.Exception e)
                {{
                    Debug.LogError($"加载{table_name}数据时发生错误: {{e.Message}}");
                    Debug.LogError($"堆栈跟踪: {{e.StackTrace}}");
                    if ({to_camel_case(table_name)}Json.text.Length > 0)
                    {{
                        Debug.LogError($"JSON前500字符: {{ {to_camel_case(table_name)}Json.text.Substring(0, Math.Min(500, {to_camel_case(table_name)}Json.text.Length)) }}");
                    }}
                }}
            }}
            else
            {{
                Debug.LogWarning($"未找到{table_name}数据文件或文件为空: ExcelImporter/{table_name}");
            }}
"""
    
    code += """
            _isInitialized = true;
            Debug.Log("ExcelDataManager初始化完成");
        }
        
        /// <summary>
        /// 获取对象的ID值（支持Id、ID、id字段，以及所有可能的变体）
        /// </summary>
        private static int? GetIdValue(object obj)
        {
            if (obj == null) return null;
            
            var type = obj.GetType();
            // Unity JsonUtility 反序列化后是字段，不是属性，所以使用 GetField
            // 尝试多种可能的ID字段名
            var idFieldNames = new[] { "Id", "ID", "id", "iD", "ID_", "Id_", "id_" };
            
            foreach (var fieldName in idFieldNames)
            {
                var idField = type.GetField(fieldName, FieldFlags);
                if (idField != null)
                {
                    var value = idField.GetValue(obj);
                    if (value is int intValue && intValue != 0)
                        return intValue;
                    if (value != null && int.TryParse(value.ToString(), out int parsedValue) && parsedValue != 0)
                        return parsedValue;
                }
            }
            
            // 如果标准ID字段都没找到，尝试查找所有字段，看是否有包含"id"的字段
            var allFields = type.GetFields(FieldFlags);
            foreach (var field in allFields)
            {
                var fieldName = field.Name;
                if (fieldName.Equals("Id", StringComparison.OrdinalIgnoreCase) || 
                    fieldName.ToLower().Contains("id"))
                {
                    var value = field.GetValue(obj);
                    if (value is int intValue && intValue != 0)
                        return intValue;
                    if (value != null && int.TryParse(value.ToString(), out int parsedValue) && parsedValue != 0)
                        return parsedValue;
                }
            }
            
            return null;
        }
"""
    
    # 为每个表生成查询方法
    for table_name, table_info in tables.items():
        class_name = to_pascal_case(table_name)
        dict_name = f"_{to_camel_case(table_name)}Dict"
        list_name = f"_{to_camel_case(table_name)}List"
        method_name = to_pascal_case(table_name)
        
        code += f"""
        /// <summary>
        /// 通过ID获取{table_name}数据
        /// </summary>
        public static {class_name}Data Get{table_name}ById(int id)
        {{
            if (!_isInitialized) Initialize();
            {dict_name}.TryGetValue(id, out var data);
            return data;
        }}
        
        /// <summary>
        /// 获取所有{table_name}数据
        /// </summary>
        public static List<{class_name}Data> GetAll{table_name}()
        {{
            if (!_isInitialized) Initialize();
            return {list_name};
        }}
        
        /// <summary>
        /// 查询{table_name}数据（支持字段查询）
        /// </summary>
        public static List<{class_name}Data> Query{table_name}(Func<{class_name}Data, bool> predicate)
        {{
            if (!_isInitialized) Initialize();
            return {list_name}.Where(predicate).ToList();
        }}
"""
    
    code += """    }
}
"""
    
    return code


def generate_array_wrapper_classes(tables: Dict[str, Dict[str, Any]]) -> str:
    """生成数组包装类代码（用于JSON反序列化）"""
    code = """using System;

namespace ExcelImporter
{
"""
    
    for table_name in tables.keys():
        class_name = to_pascal_case(table_name)
        code += f"""
    [Serializable]
    public class {class_name}DataArray
    {{
        public {class_name}Data[] items;
    }}
"""
    
    code += "}\n"
    
    return code


def to_pascal_case(name: str) -> str:
    """转换为PascalCase"""
    # 移除特殊字符，分割单词
    parts = []
    current = ""
    for char in name:
        if char.isalnum():
            current += char
        else:
            if current:
                parts.append(current)
                current = ""
    if current:
        parts.append(current)
    
    if not parts:
        parts = [name]
    
    return "".join(word.capitalize() for word in parts)


def to_camel_case(name: str) -> str:
    """转换为camelCase"""
    pascal = to_pascal_case(name)
    if not pascal:
        return pascal
    return pascal[0].lower() + pascal[1:]


def main():
    """主函数"""
    print("=" * 60)
    print("Excel数据导出工具")
    print("=" * 60)
    
    # 确保目录存在
    ensure_directories()
    
    # 检查DataTables目录
    if not DATATABLES_DIR.exists():
        print(f"错误: DataTables目录不存在: {DATATABLES_DIR}")
        return
    
    # 查找所有.xlsx文件
    xlsx_files = list(DATATABLES_DIR.glob("*.xlsx"))
    
    if not xlsx_files:
        print(f"警告: 在{DATATABLES_DIR}目录下没有找到.xlsx文件")
        return
    
    print(f"找到 {len(xlsx_files)} 个Excel文件")
    print()
    
    all_tables = {}
    
    # 处理每个Excel文件
    for xlsx_file in xlsx_files:
        print(f"处理文件: {xlsx_file.name}")
        
        tables = parse_excel_file(xlsx_file)
        if not tables:
            print(f"  跳过: 文件为空或无法解析")
            continue
        
        # 处理每个工作表
        for sheet_name, table_info in tables.items():
            table_key = f"{xlsx_file.stem}_{sheet_name}" if len(tables) > 1 else xlsx_file.stem
            
            print(f"  处理工作表: {sheet_name} -> {table_key}")
            
            # 保存JSON数据（Unity JsonUtility格式）
            json_path = UNITY_RESOURCES_DIR / f"{table_key}.json"
            # Unity的JsonUtility需要包装在对象中
            # 重要：JSON字段名必须与C#类字段名完全匹配（PascalCase）
            # 所以需要将JSON中的字段名也转换为PascalCase
            # 同时需要根据C#字段类型转换值类型
            converted_data = []
            
            # 使用与生成C#类相同的类型推断逻辑
            # 基于所有数据行推断类型（与generate_csharp_class中的逻辑一致）
            field_types_map = {}
            for header in table_info["headers"]:
                inferred_type = "string"  # 默认类型
                for row in table_info["data"]:
                    value = row.get(header)
                    if value is not None and str(value).strip():
                        inferred_type = to_csharp_type(header, value)
                        # 如果推断为int，检查是否有float值
                        if inferred_type == "int":
                            for check_row in table_info["data"]:
                                check_value = check_row.get(header)
                                if check_value is not None:
                                    try:
                                        float(str(check_value))
                                        if "." in str(check_value):
                                            inferred_type = "float"
                                            break
                                    except:
                                        pass
                        break
                field_types_map[header] = inferred_type
            
            for row in table_info["data"]:
                converted_row = {}
                for header, value in row.items():
                    # 将字段名转换为PascalCase，与C#类字段名一致
                    pascal_header = to_pascal_case(header)
                    csharp_type = field_types_map.get(header, "string")
                    
                    # 处理null值和类型转换
                    if value is None:
                        # 根据C#类型设置默认值
                        if csharp_type == "int":
                            converted_row[pascal_header] = 0
                        elif csharp_type == "float":
                            converted_row[pascal_header] = 0.0
                        elif csharp_type == "bool":
                            converted_row[pascal_header] = False
                        else:
                            converted_row[pascal_header] = ""
                    else:
                        # 确保类型匹配
                        if csharp_type == "int":
                            try:
                                converted_row[pascal_header] = int(value)
                            except (ValueError, TypeError):
                                converted_row[pascal_header] = 0
                        elif csharp_type == "float":
                            try:
                                converted_row[pascal_header] = float(value)
                            except (ValueError, TypeError):
                                converted_row[pascal_header] = 0.0
                        elif csharp_type == "bool":
                            if isinstance(value, bool):
                                converted_row[pascal_header] = value
                            else:
                                converted_row[pascal_header] = str(value).lower() in ('true', '1', 'yes')
                        else:
                            # string类型，确保是字符串
                            converted_row[pascal_header] = str(value) if value is not None else ""
                converted_data.append(converted_row)
            
            # Unity JsonUtility要求JSON格式必须严格匹配
            # 注意：JsonUtility不支持数组直接反序列化，必须包装在类中
            # Unity JsonUtility要求JSON格式必须严格匹配
            # 注意：Unity JsonUtility对数组反序列化有已知问题
            # 我们需要确保JSON格式完全正确
            json_data = {"items": converted_data}
            
            # 生成JSON字符串，确保格式正确
            # 使用ensure_ascii=False保持中文，但确保格式严格
            json_str = json.dumps(json_data, ensure_ascii=False, indent=2)
            
            # 验证JSON格式（可选，用于调试）
            try:
                json.loads(json_str)  # 验证JSON是否有效
            except json.JSONDecodeError as e:
                print(f"    警告: JSON格式验证失败: {e}")
            
            # 保存JSON文件
            with open(json_path, 'w', encoding='utf-8') as f:
                f.write(json_str)
            print(f"    已保存JSON: {json_path} (共 {len(converted_data)} 条记录)")
            
            # 生成C#类代码
            if table_info["data"]:
                class_code = generate_csharp_class(table_key, table_info["headers"], table_info["data"])
                class_path = UNITY_CODE_DIR / f"{to_pascal_case(table_key)}Data.cs"
                with open(class_path, 'w', encoding='utf-8') as f:
                    f.write(class_code)
                print(f"    已生成C#类: {class_path}")
            
            # 保存表信息
            all_tables[table_key] = table_info
    
    # 生成数据管理器代码
    if all_tables:
        print()
        print("生成数据管理器代码...")
        
        manager_code = generate_data_manager_code(all_tables)
        manager_path = UNITY_CODE_DIR / "ExcelDataManager.cs"
        with open(manager_path, 'w', encoding='utf-8') as f:
            f.write(manager_code)
        print(f"  已生成数据管理器: {manager_path}")
        
        # 生成数组包装类
        array_code = generate_array_wrapper_classes(all_tables)
        array_path = UNITY_CODE_DIR / "ExcelDataArray.cs"
        with open(array_path, 'w', encoding='utf-8') as f:
            f.write(array_code)
        print(f"  已生成数组包装类: {array_path}")
    
    print()
    print("=" * 60)
    print("导出完成！")
    print("=" * 60)
    print(f"数据文件: {UNITY_RESOURCES_DIR}")
    print(f"C#代码: {UNITY_CODE_DIR}")
    print()
    print("使用方法:")
    print("  在Unity中调用: ExcelDataManager.Initialize()")
    print("  然后使用: ExcelDataManager.Get表名ById(id) 或 ExcelDataManager.Query表名(predicate)")


if __name__ == "__main__":
    main()

